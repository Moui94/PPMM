"""
backend/routes/export.py
Excel-Export-API.

GET /api/export/full              → Alle aktiven Aufträge als Excel
GET /api/export/order/<pa_nr>     → Einzelauftrag als Excel
"""

import io
from datetime import date
from flask import Blueprint, send_file, request
from backend.database import get_db
from backend.models.order import list_orders, get_order_by_pa_nr
from backend.models.operation import get_operations
from backend.services.date_calc import fmt_date_ch
from ._helpers import error, login_required, role_required

export_bp = Blueprint("export", __name__, url_prefix="/api/export")


def _build_workbook(orders_with_ops: list[tuple]) -> io.BytesIO:
    """
    Erstellt ein openpyxl-Workbook mit einer Zeile pro Auftrag + AGs.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl nicht installiert. Bitte: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aufträge"

    # Statische Spalten
    static_cols = [
        "PA-Nr", "Typ", "Artikel", "Ceramaret", "Spezielles",
        "Menge", "Priorität", "Status", "PA-Start",
        "Endtermin Soll", "Auslieferung Kunde", "Abweichung (T)", "Bemerkung",
    ]
    # Dynamische AG-Spalten (AG01 Start/Soll/Ende/Maschine)
    ag_cols = []
    for ag_nr in range(1, 15):
        ag_cols += [
            f"AG{ag_nr:02d} Start", f"AG{ag_nr:02d} Dauer",
            f"AG{ag_nr:02d} Ende", f"AG{ag_nr:02d} Maschine",
        ]

    all_cols = static_cols + ag_cols
    ws.append(all_cols)

    # Header-Styling
    header_fill = PatternFill("solid", fgColor="1E2A38")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 14

    # Zeilen füllen
    for order, ops in orders_with_ops:
        # AG-Dict aufbauen
        ops_by_nr = {op["ag_nr"]: op for op in ops}

        row = [
            order["pa_nr"],
            "Implantat" if order["art"] == "I" else "Abutment",
            order["artikel"],
            order["ceramaret"] or "",
            order["spezielles"] or "",
            order["menge"],
            order["prioritaet"],
            order["status"],
            fmt_date_ch(order["pa_start"]),
            fmt_date_ch(order["endtermin_soll"]),
            fmt_date_ch(order["auslieferung_kunde"]),
            order["abweichung_tage"] or 0,
            order["bemerkung"] or "",
        ]
        for ag_nr in range(1, 15):
            op = ops_by_nr.get(ag_nr)
            if op:
                row += [
                    fmt_date_ch(op["start_soll"]),
                    op["solldauer_tage"],
                    fmt_date_ch(op["ende_soll"]),
                    op["maschine"] or "",
                ]
            else:
                row += ["", "", "", ""]

        ws.append(row)

        # Verzug → Zeile rot färben
        if (order["abweichung_tage"] or 0) > 0:
            red = PatternFill("solid", fgColor="FFE0E0")
            for cell in ws[ws.max_row]:
                cell.fill = red

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@export_bp.route("/full", methods=["GET"])
@login_required
def export_full():
    conn      = get_db()
    status    = request.args.get("status", "aktiv")
    orders    = list_orders(conn, status=status if status != "alle" else None)
    with_ops  = [(o, get_operations(conn, o["id"])) for o in orders]

    try:
        buf = _build_workbook(with_ops)
    except RuntimeError as e:
        return error(str(e), 500)

    filename = f"auftraege_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@export_bp.route("/order/<pa_nr>", methods=["GET"])
@login_required
def export_order(pa_nr):
    conn  = get_db()
    order = get_order_by_pa_nr(conn, pa_nr)
    if not order:
        return error(f"Auftrag '{pa_nr}' nicht gefunden.", 404)
    ops     = get_operations(conn, order["id"])
    try:
        buf = _build_workbook([(order, ops)])
    except RuntimeError as e:
        return error(str(e), 500)

    filename = f"auftrag_{pa_nr}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
